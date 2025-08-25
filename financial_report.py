from bs4 import BeautifulSoup
import re
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font
import os
import shutil
from datetime import datetime

def extract_dates(html_file):
    with open(html_file, 'r', encoding='windows-1250') as f:
        content = f.read()
    match = re.search(r'Za razdoblje \(po datumu valute\): (\d{2}\.\d{2}\.\d{4})\. do (\d{2}\.\d{2}\.\d{4})\.', content)
    return (match.group(1), match.group(2)) if match else (None, None)

def get_first_day_of_month(date_str):
    # Extract just the date part if there's time info
    date_part = date_str.split()[0] if ' ' in date_str else date_str
    parts = date_part.split('.')
    if len(parts) >= 3:
        day, month, year = parts[0], parts[1], parts[2]
        return f"01.{month}.{year}"
    return date_str

def date_to_datetime(date_str):
    day, month, year = date_str.split('.')
    return datetime(int(year), int(month), int(day))

def spans_multiple_months(start_date, end_date):
    start_dt = date_to_datetime(start_date)
    end_dt = date_to_datetime(end_date)
    return start_dt.month != end_dt.month or start_dt.year != end_dt.year

def check_period_overlap(start_date, end_date, log_file):
    periods_file = 'processed_periods'
    if os.path.exists(periods_file):
        start_dt = date_to_datetime(start_date)
        end_dt = date_to_datetime(end_date)
        
        with open(periods_file, 'r') as f:
            for line in f:
                if line.strip():
                    existing_start, existing_end = line.strip().split(',')
                    existing_start_dt = date_to_datetime(existing_start)
                    existing_end_dt = date_to_datetime(existing_end)
                    
                    # Check for actual date overlap
                    if start_dt <= existing_end_dt and end_dt >= existing_start_dt:
                        with open(log_file, 'a', encoding='utf-8') as lf:
                            lf.write(f"OVERLAP DETECTED: Period {start_date}-{end_date} overlaps with existing {existing_start}-{existing_end}\n")
                            lf.write(f"ACTION: Skipping processing to prevent duplicate transactions\n")
                        print(f"  Overlap with existing period: {existing_start} to {existing_end}")
                        return True
    
    with open(periods_file, 'a') as f:
        f.write(f"{start_date},{end_date}\n")
    print(f"  Period registered: {start_date} to {end_date}")
    return False

def log_transaction(log_file, description, amount, category, skipped=False):
    timestamp = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
    status = "SKIPPED" if skipped else "PROCESSED"
    with open(log_file, 'a', encoding='utf-8') as f:
        f.write(f"{timestamp} | {description} | {amount} | {category} | {status}\n")

def parse_transactions(html_file, log_file):
    with open(html_file, 'r', encoding='windows-1250') as f:
        content = f.read()
    
    soup = BeautifulSoup(content, 'html.parser')
    table = soup.find('table', {'style': lambda x: x and 'border-collapse: collapse' in x})
    transactions = []
    
    with open(log_file, 'w', encoding='utf-8') as f:
        f.write("Timestamp | Description | Amount | Category | Status\n")
        f.write("-" * 80 + "\n")
    
    for row in table.find_all('tr'):
        cells = row.find_all('td')
        if len(cells) >= 6 and not row.get('bgcolor'):
            date_text = cells[0].get_text(strip=True)
            if re.match(r'\d{2}\.\d{2}\.\d{4}', date_text):
                # Extract first date, handle both \n and <br> separators
                transaction_date = re.search(r'(\d{2}\.\d{2}\.\d{4})', date_text).group(1)
                description = cells[1].get_text(strip=True).split('\n')[0].strip()
                amount_text = cells[4].get_text(strip=True)
                if amount_text and amount_text != '&nbsp;':
                    try:
                        amount = float(amount_text.replace(',', '.').replace(' ', ''))
                        transactions.append({'date': transaction_date, 'description': description, 'amount': amount})
                    except ValueError:
                        log_transaction(log_file, description, amount_text, "N/A", skipped=True)
                else:
                    log_transaction(log_file, description, "0.00", "N/A", skipped=True)
    return transactions

def load_mapping():
    # Use category_mapping.csv, not the example file
    mapping_file = 'category_mapping.csv'
    if os.path.exists(mapping_file) and mapping_file != 'category_mapping_example.csv':
        mapping = {}
        with open(mapping_file, 'r', encoding='utf-8') as f:
            lines = f.readlines()[1:]
            for line in lines:
                parts = line.strip().split(',')
                if len(parts) >= 2:
                    mapping[parts[0]] = parts[1]
        return mapping
    return {}

def categorize(description, mapping):
    for keyword, category in mapping.items():
        if keyword.lower() in description.lower():
            return category
    return 'Other'

def save_excel(totals, transactions, end_date):
    sheet_name = get_first_day_of_month(end_date)
    month_file = f"financial_report_{sheet_name.replace('.', '-')}.xlsx"
    print(f"Saving data to file: {month_file}")
    
    changes = []
    file_existed = os.path.exists(month_file)
    
    if file_existed:
        backup_date = datetime.now().strftime("%d_%m_%Y")
        shutil.copy2(month_file, f"{month_file}.backup_{backup_date}")
        wb = load_workbook(month_file)
    else:
        wb = Workbook()
        wb.remove(wb.active)
        changes.append(f"Created new file: {month_file}")
    
    # Create sheets in order: details, cash, totals
    if 'details' in wb.sheetnames:
        ws_details = wb['details']
    else:
        ws_details = wb.create_sheet('details')
        changes.append("Created details sheet")
    
    if 'cash' in wb.sheetnames:
        ws_cash = wb['cash']
    else:
        ws_cash = wb.create_sheet('cash')
        changes.append("Created cash sheet")
    
    if 'totals' in wb.sheetnames:
        ws_totals = wb['totals']
    else:
        ws_totals = wb.create_sheet('totals')
        changes.append("Created totals sheet")
    
    bold = Font(bold=True)
    
    # Get budget values from environment variables
    income = float(os.getenv('FIN_INCOME_VALUE', '6453'))
    taxes = float(os.getenv('FIN_TAX_VALUE', '1652'))
    food_budget = float(os.getenv('FIN_FOOD_VALUE', '900'))
    utility_budget = float(os.getenv('FIN_UTILITY_VALUE', '1333.88'))
    savings = float(os.getenv('FIN_SAVINGS_VALUE', '700'))
    
    # Calculate budget for other category
    other_budget = income - taxes - food_budget - utility_budget - savings
    
    # 1. Setup details sheet
    if not ws_details['A1'].value:
        ws_details['A1'] = 'Дата'
        ws_details['A1'].font = bold
        ws_details['B1'] = 'Категория'
        ws_details['B1'].font = bold
        ws_details['C1'] = 'Сумма'
        ws_details['C1'].font = bold
    
    # Find next empty row to append transactions
    row = 2
    while ws_details[f'A{row}'].value:
        row += 1
    
    # Add transaction details
    for transaction in transactions:
        ws_details[f'A{row}'] = transaction['date']
        ws_details[f'B{row}'] = transaction['category']
        ws_details[f'C{row}'] = transaction['amount']
        row += 1
        changes.append(f"Added transaction: {transaction['date']} - {transaction['category']} - {transaction['amount']:.2f}")
    
    # 2. Setup cash sheet headers (only if not already set)
    if not ws_cash['A1'].value:
        ws_cash['A1'] = 'Дата'
        ws_cash['A1'].font = bold
        ws_cash['B1'] = 'Категория'
        ws_cash['B1'].font = bold
        ws_cash['C1'] = 'Сумма'
        ws_cash['C1'].font = bold
    
    # 3. Calculate totals from details and cash sheets
    def get_category_total(sheet, category):
        total = 0
        row = 2
        while sheet[f'B{row}'].value:
            if sheet[f'B{row}'].value == category and sheet[f'C{row}'].value:
                total += float(sheet[f'C{row}'].value)
            row += 1
        return total
    
    def get_cash_total(sheet):
        total = 0
        row = 2
        while sheet[f'C{row}'].value:
            total += float(sheet[f'C{row}'].value)
            row += 1
        return total
    
    details_food = get_category_total(ws_details, 'Food')
    details_utility = get_category_total(ws_details, 'Utility bills')
    details_other = get_category_total(ws_details, 'Other')
    
    cash_food = get_category_total(ws_cash, 'Food')
    cash_utility = get_category_total(ws_cash, 'Utility bills')
    cash_other = get_category_total(ws_cash, 'Other')
    cash_total = get_cash_total(ws_cash)
    
    # 4. Setup totals sheet with new layout
    ws_totals['A1'] = 'Доход'
    ws_totals['A1'].font = bold
    if not ws_totals['A2'].value:
        ws_totals['A2'] = income
    
    ws_totals['B1'] = 'Налоги'
    ws_totals['B1'].font = bold
    if not ws_totals['B2'].value:
        ws_totals['B2'] = taxes
    
    ws_totals['C1'] = 'Еда'
    ws_totals['C1'].font = bold
    if not ws_totals['C2'].value:
        ws_totals['C2'] = food_budget
    ws_totals['C3'] = food_budget - details_food - cash_food
    
    ws_totals['D1'] = 'Комы'
    ws_totals['D1'].font = bold
    if not ws_totals['D2'].value:
        ws_totals['D2'] = utility_budget
    ws_totals['D3'] = utility_budget - details_utility - cash_utility
    
    ws_totals['E1'] = 'Отложить'
    ws_totals['E1'].font = bold
    if not ws_totals['E2'].value:
        ws_totals['E2'] = savings
    
    ws_totals['F1'] = 'Бюджет'
    ws_totals['F1'].font = bold
    if not ws_totals['F2'].value:
        ws_totals['F2'] = other_budget
    ws_totals['F3'] = other_budget - details_other - cash_other
    
    ws_totals['G1'] = 'Наличка'
    ws_totals['G1'].font = bold
    ws_totals['G2'] = cash_total
    
    ws_totals['H1'] = 'Заполнено по'
    ws_totals['H1'].font = bold
    ws_totals['H2'] = end_date
    

    
    wb.save(month_file)
    return changes, month_file

def main():
    # Check if running in Docker (data directory exists)
    if os.path.exists('/app/data'):
        os.chdir('/app/data')
    
    # Find all HTML files in current directory, excluding examples
    html_files = [f for f in os.listdir('.') if f.endswith('.html') and f != 'statement_example.html']
    
    if not html_files:
        print("Error: No HTML files found")
        return
    
    print(f"Found {len(html_files)} HTML files: {html_files}")
    
    mapping = load_mapping()
    total_processed = 0
    
    for html_file in html_files:
        print(f"\n=== Processing {html_file} ===")
        
        start_date, end_date = extract_dates(html_file)
        if not start_date or not end_date:
            print(f"ERROR: Could not extract dates from {html_file}")
            continue
        
        # Check if period spans multiple months
        if spans_multiple_months(start_date, end_date):
            print(f"WARNING: File spans multiple months ({start_date} to {end_date})")
            print(f"SKIPPED: Multi-month periods are not supported")
            continue
        
        sheet_name = get_first_day_of_month(end_date)
        print(f"Statement period: {start_date} to {end_date}")
        print(f"Target Excel sheet: {sheet_name}")
        
        log_file = f"logs_{end_date.replace('.', '_')}.log"
        
        if check_period_overlap(start_date, end_date, log_file):
            print(f"SKIPPED: Period overlap detected. Check {log_file}")
            print(f"Reason: Period {start_date}-{end_date} overlaps with existing processed period")
            continue
        
        transactions = parse_transactions(html_file, log_file)
        print(f"Found {len(transactions)} transactions in {html_file}")
        
        # Process all transactions for this statement period
        totals = {}
        for t in transactions:
            category = categorize(t['description'], mapping)
            totals[category] = totals.get(category, 0) + t['amount']
            log_transaction(log_file, t['description'], f"€{t['amount']:.2f}", category)
        
        # Prepare transaction details for details sheet
        transaction_details = []
        for t in transactions:
            category = categorize(t['description'], mapping)
            transaction_details.append({
                'date': t['date'],
                'category': category,
                'amount': t['amount']
            })
        
        # Save to monthly file
        changes, month_file = save_excel(totals, transaction_details, end_date)
        print(f"SUCCESS: Created/updated file '{month_file}' with {len(transactions)} transactions")
        for category, total in totals.items():
            print(f"  {category}: €{total:.2f}")
        
        if changes:
            print(f"  Changes made:")
            # Show only non-transaction changes to avoid spam
            non_transaction_changes = [c for c in changes if not c.startswith('Added transaction:')]
            transaction_count = len([c for c in changes if c.startswith('Added transaction:')])
            
            for change in non_transaction_changes:
                print(f"    - {change}")
            if transaction_count > 0:
                print(f"    - Added {transaction_count} transactions to details sheet")
        
        total_processed += len(transactions)
    
    print(f"\n=== SUMMARY ===")
    print(f"Total processed: {total_processed} transactions")
    print(f"Files saved in: {os.getcwd()}")
    
    # Show final report summary
    excel_files = [f for f in os.listdir('.') if f.startswith('financial_report_') and f.endswith('.xlsx')]
    if excel_files:
        print(f"\nCreated {len(excel_files)} monthly financial reports:")
        for file in sorted(excel_files):
            print(f"  - {file}")

if __name__ == "__main__":
    main()